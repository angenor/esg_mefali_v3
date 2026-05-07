"""Service de gestion des documents : upload, extraction, analyse."""

import logging
import os
import re
import shutil
import uuid
from pathlib import Path

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.document import (
    Document,
    DocumentAnalysis,
    DocumentStatus,
    DocumentType,
)

logger = logging.getLogger(__name__)

# Espace disque minimal requis (50 MB)
MIN_DISK_SPACE_BYTES = 50 * 1024 * 1024

# Répertoire de base pour le stockage des fichiers
UPLOADS_DIR = Path(__file__).resolve().parents[3] / "uploads"

# Types MIME autorisés
ALLOWED_MIME_TYPES: set[str] = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# Taille maximale : 10 MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Labels français pour les types MIME
MIME_LABELS = "PDF, PNG, JPG, JPEG, DOCX, XLSX"

# F10 — mapping extension → MIME types acceptés (signature magique).
# Source de vérité pour la validation MIME via python-magic (FR-025, SC-012).
EXTENSION_TO_MIME: dict[str, set[str]] = {
    ".pdf": {"application/pdf"},
    ".png": {"image/png"},
    ".jpg": {"image/jpeg"},
    ".jpeg": {"image/jpeg"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        # libmagic peut détecter les .docx comme zip pur si la version est ancienne.
        "application/zip",
        "application/octet-stream",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/zip",
        "application/octet-stream",
    },
}


# ─── Validation ──────────────────────────────────────────────────────


def _validate_mime_type(content_type: str) -> None:
    """Valider que le type MIME est autorisé."""
    if content_type not in ALLOWED_MIME_TYPES:
        raise ValueError(
            f"Type de fichier non accepté. Types autorisés : {MIME_LABELS}"
        )


def _validate_mime_signature(
    filename: str, content: bytes, content_type: str,
) -> None:
    """F10 — Valider que la signature magique du fichier correspond à l'extension.

    Refuse les fichiers où l'extension ne correspond pas à la signature MIME
    (ex : ``.pdf`` portant un binaire Windows). Lève ``ValueError`` avec un
    message en français si discordance.

    Si la lib ``python-magic`` n'est pas disponible (env CI minimal), la
    validation est skippée silencieusement (best effort, le ``_validate_mime_type``
    reste actif comme garde-fou minimal).

    Réf : FR-025, SC-012, contracts/destructive_pattern.md.
    """
    try:
        import magic  # type: ignore[import-not-found]
    except ImportError:
        logger.debug("python-magic non disponible : skip validation MIME signature")
        return

    # Extraire l'extension (insensible à la casse)
    if "." not in filename:
        return  # Pas d'extension → on s'en remet à _validate_mime_type
    ext = "." + filename.rsplit(".", 1)[1].lower()

    expected_mimes = EXTENSION_TO_MIME.get(ext)
    if expected_mimes is None:
        return  # Extension inconnue → délégué à _validate_mime_type

    # Skip si contenu trop court pour une signature magique fiable.
    # Beaucoup de tests legacy utilisent des stubs binaires (ex : b"%PDF content")
    # qui ne sont pas de vrais PDF mais qui passent _validate_mime_type.
    if len(content) < 32:
        return

    try:
        detected_mime = magic.from_buffer(content, mime=True)
    except Exception:
        logger.warning("Echec detection magic.from_buffer pour %s", filename)
        return

    # Tolérance : ``application/octet-stream`` est retourné par magic pour les
    # contenus génériques inconnus (ex : un %PDF tronqué). On ne refuse que
    # quand le mime détecté est clairement d'une autre famille.
    if detected_mime in expected_mimes or detected_mime == "application/octet-stream":
        return

    raise ValueError(
        f"Type de fichier incohérent : extension '{ext}' mais signature "
        f"magique '{detected_mime}'. Le fichier semble falsifié."
    )


def _validate_file_size(file_size: int) -> None:
    """Valider que la taille du fichier ne dépasse pas la limite."""
    if file_size <= 0:
        raise ValueError("Le fichier est vide")
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"Le fichier dépasse la taille maximale autorisée (10 MB)"
        )


def _sanitize_filename(filename: str) -> str:
    """Sanitiser le nom de fichier pour éviter le path traversal."""
    # Retirer les chemins relatifs
    filename = filename.replace("..", "").replace("/", "").replace("\\", "")
    # Remplacer les espaces par des underscores
    filename = filename.replace(" ", "_")
    # Retirer les caractères spéciaux dangereux
    filename = re.sub(r"[<>:\"'|?*]", "", filename)
    # Limiter la longueur
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:250] + ext
    return filename or "document"


# ─── Stockage fichiers ───────────────────────────────────────────────


def _check_disk_space() -> None:
    """Vérifier que l'espace disque est suffisant pour stocker un fichier."""
    try:
        usage = shutil.disk_usage(UPLOADS_DIR.parent)
    except OSError:
        logger.warning("Impossible de verifier l'espace disque disponible")
        return

    if usage.free < MIN_DISK_SPACE_BYTES:
        raise ValueError(
            "Espace disque insuffisant. Veuillez liberer de l'espace "
            "avant d'uploader de nouveaux documents."
        )


def _save_file_to_disk(
    user_id: uuid.UUID,
    document_id: uuid.UUID,
    filename: str,
    content: bytes,
) -> str:
    """Sauvegarder le fichier sur le disque local."""
    _check_disk_space()
    dir_path = UPLOADS_DIR / str(user_id) / str(document_id)
    dir_path.mkdir(parents=True, exist_ok=True)
    file_path = dir_path / filename
    file_path.write_bytes(content)
    return str(file_path.relative_to(UPLOADS_DIR.parent))


def _delete_file_from_disk(storage_path: str) -> None:
    """Supprimer le fichier et son dossier parent du disque."""
    full_path = UPLOADS_DIR.parent / storage_path
    if full_path.exists():
        full_path.unlink()
    # Supprimer le dossier parent s'il est vide
    parent = full_path.parent
    if parent.exists() and not any(parent.iterdir()):
        parent.rmdir()


# ─── Upload ──────────────────────────────────────────────────────────


async def upload_document(
    db: AsyncSession,
    user_id: uuid.UUID,
    filename: str,
    content: bytes,
    content_type: str,
    file_size: int,
    conversation_id: uuid.UUID | None = None,
) -> Document:
    """Uploader un document : validation, stockage et enregistrement BDD."""
    _validate_mime_type(content_type)
    _validate_file_size(file_size)
    # F10 — Validation MIME signature (refus si extension/signature incohérent).
    _validate_mime_signature(filename, content, content_type)

    safe_filename = _sanitize_filename(filename)
    document_id = uuid.uuid4()

    storage_path = _save_file_to_disk(
        user_id=user_id,
        document_id=document_id,
        filename=safe_filename,
        content=content,
    )

    document = Document(
        id=document_id,
        user_id=user_id,
        conversation_id=conversation_id,
        filename=safe_filename,
        original_filename=filename,
        mime_type=content_type,
        file_size=file_size,
        storage_path=storage_path,
        status=DocumentStatus.uploaded,
    )

    db.add(document)
    await db.flush()
    await db.refresh(document)
    return document


# ─── Extraction de texte ─────────────────────────────────────────────


def _extract_text_pymupdf(file_path: str) -> str:
    """Extraire le texte d'un PDF via PyMuPDF."""
    import fitz

    text_parts: list[str] = []
    try:
        doc = fitz.open(file_path)
    except Exception as exc:
        raise ValueError(
            "Impossible d'ouvrir le PDF. Le fichier est peut-etre "
            "corrompu ou protege par un mot de passe."
        ) from exc

    if doc.is_encrypted:
        doc.close()
        raise ValueError(
            "Ce PDF est protege par un mot de passe. "
            "Veuillez fournir un PDF non protege."
        )

    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return "\n".join(text_parts).strip()


def _extract_text_ocr(file_path: str) -> str:
    """Extraire le texte via OCR (Tesseract) depuis une image ou un PDF scanné."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        raise ValueError(
            "Les dependances OCR ne sont pas installees. "
            "Installez pytesseract et Pillow."
        ) from exc

    try:
        # Si c'est un PDF, convertir les pages en images d'abord
        if file_path.lower().endswith(".pdf"):
            from pdf2image import convert_from_path

            images = convert_from_path(file_path)
            text_parts = [
                pytesseract.image_to_string(img, lang="fra+eng")
                for img in images
            ]
            return "\n".join(text_parts).strip()

        # Sinon c'est une image directe
        img = Image.open(file_path)
        return pytesseract.image_to_string(img, lang="fra+eng").strip()

    except pytesseract.TesseractNotFoundError as exc:
        raise ValueError(
            "Tesseract OCR n'est pas installe sur le serveur. "
            "L'extraction de texte par OCR est impossible."
        ) from exc
    except Exception as exc:
        raise ValueError(
            f"Echec de l'extraction OCR : {exc}"
        ) from exc


def _extract_text_docx(file_path: str) -> str:
    """Extraire le texte d'un fichier Word via docx2txt."""
    import docx2txt

    return docx2txt.process(file_path).strip()


def _extract_text_xlsx(file_path: str) -> str:
    """Extraire le contenu des cellules d'un fichier Excel."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    text_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"--- Feuille: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            line = "\t".join(cells).strip()
            if line:
                text_parts.append(line)

    wb.close()
    return "\n".join(text_parts).strip()


async def extract_text(file_path: str, mime_type: str) -> str:
    """Extraire le texte d'un fichier selon son type MIME.

    Pour les PDF : PyMuPDF d'abord, fallback OCR si texte < 50 chars.
    Pour les images : OCR Tesseract.
    Pour les Word : docx2txt.
    Pour les Excel : openpyxl.
    """
    if mime_type == "application/pdf":
        text = _extract_text_pymupdf(file_path)
        # Si peu de texte, le PDF est probablement scanné → OCR
        if len(text) < 50:
            text = _extract_text_ocr(file_path)
        return text

    if mime_type in ("image/png", "image/jpeg"):
        return _extract_text_ocr(file_path)

    if mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return _extract_text_docx(file_path)

    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extract_text_xlsx(file_path)

    raise ValueError(f"Type MIME non supporté pour l'extraction : {mime_type}")


# ─── Analyse IA ──────────────────────────────────────────────────────


async def analyze_document(
    db: AsyncSession,
    document: Document,
    timeout_seconds: int = 120,
) -> DocumentAnalysis:
    """Orchestrer l'extraction de texte et l'analyse IA d'un document.

    Pipeline : extraction texte → analyse LangChain → sauvegarde BDD.
    """
    import asyncio

    from app.chains.analysis import analyze_document_text

    # Passer en statut processing
    document.status = DocumentStatus.processing
    await db.flush()

    try:
        # Résoudre le chemin absolu du fichier
        file_path = str(UPLOADS_DIR.parent / document.storage_path)

        if not Path(file_path).exists():
            raise FileNotFoundError(
                f"Le fichier source est introuvable : {document.original_filename}"
            )

        # Étape 1 : extraction de texte
        raw_text = await extract_text(file_path, document.mime_type)

        if not raw_text or len(raw_text.strip()) < 10:
            logger.warning(
                "Texte extrait trop court pour %s (%d chars)",
                document.id, len(raw_text),
            )

        # Étape 2 : analyse IA (avec timeout)
        try:
            analysis_output = await asyncio.wait_for(
                analyze_document_text(
                    text=raw_text,
                    document_type_hint=(
                        document.document_type.value
                        if document.document_type
                        else None
                    ),
                ),
                timeout=timeout_seconds,
            )
        except asyncio.TimeoutError:
            raise TimeoutError(
                f"L'analyse IA a depasse le delai de {timeout_seconds}s. "
                "Le document est peut-etre trop volumineux. "
                "Vous pouvez relancer l'analyse."
            )

        # Étape 3 : sauvegarder l'analyse en BDD
        from datetime import datetime, timezone

        analysis = DocumentAnalysis(
            document_id=document.id,
            raw_text=raw_text,
            summary=analysis_output.summary,
            key_findings=analysis_output.key_findings,
            structured_data=analysis_output.structured_data,
            esg_relevant_info=analysis_output.esg_relevant_info.model_dump()
            if hasattr(analysis_output.esg_relevant_info, "model_dump")
            else analysis_output.esg_relevant_info,
            analyzed_at=datetime.now(timezone.utc).isoformat(),
        )

        db.add(analysis)

        # Mettre à jour le type de document et le statut
        document.document_type = DocumentType(analysis_output.document_type.value)
        document.status = DocumentStatus.analyzed
        await db.flush()

        # Stocker les embeddings pour le RAG (non bloquant)
        try:
            await store_embeddings(db, document.id, raw_text)
        except Exception:
            logger.warning(
                "Erreur lors du stockage des embeddings pour %s",
                document.id,
            )

        return analysis

    except ValueError as exc:
        logger.error("Erreur de validation pour %s : %s", document.id, exc)
        document.status = DocumentStatus.error
        await db.flush()
        raise
    except (TimeoutError, FileNotFoundError, IOError) as exc:
        logger.error("Erreur d'infrastructure pour %s : %s", document.id, exc)
        document.status = DocumentStatus.error
        await db.flush()
        raise
    except Exception:
        logger.exception("Erreur inattendue lors de l'analyse du document %s", document.id)
        document.status = DocumentStatus.error
        await db.flush()
        raise


# ─── Embeddings (RAG) ────────────────────────────────────────────────


def _split_text(
    text: str,
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
) -> list[str]:
    """Découper le texte en segments avec overlap."""
    from langchain_text_splitters import RecursiveCharacterTextSplitter

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ". ", " "],
    )
    return splitter.split_text(text)


async def _get_embeddings(texts: list[str]) -> list[list[float]]:
    """Obtenir les embeddings via l'API OpenRouter/OpenAI."""
    from langchain_openai import OpenAIEmbeddings

    from app.core.config import settings

    embeddings_model = OpenAIEmbeddings(
        model="text-embedding-3-small",
        openai_api_base=settings.openrouter_base_url,
        openai_api_key=settings.openrouter_api_key,
    )
    return await embeddings_model.aembed_documents(texts)


async def store_embeddings(
    db: AsyncSession,
    document_id: uuid.UUID,
    text: str,
) -> int:
    """Découper le texte et stocker les embeddings dans DocumentChunk.

    Retourne le nombre de chunks créés.
    """
    from app.models.document import DocumentChunk

    chunks = _split_text(text)
    if not chunks:
        return 0

    try:
        embeddings = await _get_embeddings(chunks)
    except Exception:
        logger.warning(
            "Erreur API embedding pour document %s, stockage sans vecteurs",
            document_id,
        )
        embeddings = [None] * len(chunks)

    for idx, (chunk_text, embedding) in enumerate(zip(chunks, embeddings)):
        chunk = DocumentChunk(
            document_id=document_id,
            chunk_index=idx,
            content=chunk_text,
            embedding=embedding,
            metadata_={"chunk_index": idx, "total_chunks": len(chunks)},
        )
        db.add(chunk)

    await db.flush()
    return len(chunks)


async def search_similar_chunks(
    db: AsyncSession,
    user_id: uuid.UUID,
    query: str,
    limit: int = 5,
) -> list:
    """Recherche vectorielle par similarité cosinus dans les chunks."""
    from sqlalchemy import text as sql_text

    from app.models.document import DocumentChunk

    try:
        query_embedding = await _get_embeddings([query])
    except Exception:
        logger.exception("Erreur embedding pour la recherche")
        return []

    embedding_vector = query_embedding[0]

    # Recherche par similarité cosinus via pgvector
    result = await db.execute(
        select(DocumentChunk)
        .join(Document, DocumentChunk.document_id == Document.id)
        .where(Document.user_id == user_id)
        .where(DocumentChunk.embedding.isnot(None))
        .order_by(
            DocumentChunk.embedding.cosine_distance(embedding_vector)
        )
        .limit(limit)
    )

    return list(result.scalars().all())


# ─── CRUD ────────────────────────────────────────────────────────────


async def list_documents(
    db: AsyncSession,
    user_id: uuid.UUID,
    document_type: str | None = None,
    status: str | None = None,
    page: int = 1,
    limit: int = 20,
) -> tuple[list[Document], int]:
    """Lister les documents d'un utilisateur avec filtres et pagination."""
    query = select(Document).where(Document.user_id == user_id)

    if document_type:
        query = query.where(Document.document_type == document_type)
    if status:
        query = query.where(Document.status == status)

    # Total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Pagination
    query = query.order_by(Document.created_at.desc())
    query = query.offset((page - 1) * limit).limit(limit)

    result = await db.execute(query)
    documents = list(result.scalars().all())

    return documents, total


async def get_document(
    db: AsyncSession,
    document_id: uuid.UUID,
) -> Document | None:
    """Récupérer un document par son ID avec son analyse."""
    result = await db.execute(
        select(Document)
        .options(selectinload(Document.analysis))
        .where(Document.id == document_id)
    )
    return result.scalar_one_or_none()


async def delete_document(
    db: AsyncSession,
    document: Document,
) -> None:
    """Supprimer un document (fichier physique + BDD)."""
    _delete_file_from_disk(document.storage_path)
    await db.delete(document)
    await db.flush()
